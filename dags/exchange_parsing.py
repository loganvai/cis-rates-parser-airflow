import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from io import BytesIO


headers = {'User_agent':'Mozilla/5.0 (Windows   NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36'}

def get_start_end_dates():
        today = datetime.today()
        first_day = today.replace(day=1)
        end_date = first_day - timedelta(1)
        start_date = end_date.replace(day=1)
        return start_date, end_date

start_date, end_date = get_start_end_dates()

def tj_exchange(start_date, end_date):
    exchange_dict = {'date': [],
                     'exchange_rate_USD': [],
                     'exchange_rate_EUR': []}

    url = 'https://www.nbt.tj/ru/kurs/kurs.php?date='
    current_date = start_date

    # парсер
    while current_date <= end_date:
        date_str = current_date.strftime('%d.%m.%Y')
        base_url = url + date_str
        response = requests.get(base_url, headers=headers)
        soap = BeautifulSoup(response.text, 'html.parser')
        rows = soap.find_all('tr')
        usd_cells = rows[1].find_all('td')
        eur_cells = rows[2].find_all('td')
        if not eur_cells or not usd_cells:
            continue
        exchange_dict['date'].append(date_str)
        exchange_dict['exchange_rate_USD'].append(usd_cells[4].text.strip())
        exchange_dict['exchange_rate_EUR'].append(eur_cells[4].text.strip())
        current_date += timedelta(1)

    df = pd.DataFrame(exchange_dict)

    # Создаём Excel-книгу
    wb = Workbook()
    del wb['Sheet']  # удалим пустой лист

    # Добавим столбец с месяцем в формате "YYYY-MM"
    df['month'] = pd.to_datetime(df['date'], format=('%d.%m.%Y')).dt.strftime('%B, %Y')

    # Создаем ExcelWriter
    with pd.ExcelWriter('/opt/airflow/dags/tj_rates.xlsx') as writer:
        for month, group in df.groupby('month'):
            group_clean = group.drop(columns='month').copy()

            # Считаем средние значения
            avg_usd = pd.to_numeric(group_clean['exchange_rate_USD'].str.replace(',', '.'), errors='coerce').mean()
            avg_eur = pd.to_numeric(group_clean['exchange_rate_EUR'].str.replace(',', '.'), errors='coerce').mean()

            # Добавляем строку со средними
            avg_row = {
                'date': 'Среднее за месяц',
                'exchange_rate_USD': f"{avg_usd:.4f}",
                'exchange_rate_EUR': f"{avg_eur:.4f}"
            }
            group_clean = pd.concat([group_clean, pd.DataFrame([avg_row])], ignore_index=True)

            # Сохраняем в лист Excel
            group_clean.to_excel(writer, sheet_name=month, index=False)
        print('Файл Таджикистана успешно сохранен')

def kz_exchange(start_date,end_date):
    exchange_dict = {'date': [],
                     'exchange_rate_USD': [],
                     'exchange_rate_EUR': []}

    url = f'https://nationalbank.kz/ru/exchangerates/ezhednevnye-oficialnye-rynochnye-kursy-valyut/report?beginDate={start_date.strftime('%d.%m.%Y')}&endDate={end_date.strftime('%d.%m.%Y')}&search-exchanges=&rates%5B%5D=5&rates%5B%5D=6'
    response = requests.get(url,headers=headers)
    soap = BeautifulSoup(response.text, 'html.parser')
    rows = soap.find_all('tr')
    for row in rows:
        cells = row.find_all('td')
        if not cells:
            continue
        exchange_dict['date'].append(cells[0].text.strip())
        exchange_dict['exchange_rate_USD'].append(cells[2].text.strip())
        exchange_dict['exchange_rate_EUR'].append(cells[4].text.strip())
        df = pd.DataFrame(exchange_dict)
        # Создаём Excel-книгу
    wb = Workbook()
    del wb['Sheet']  # удалим пустой лист

    # Добавим столбец с месяцем в формате "YYYY-MM"
    df['month'] = pd.to_datetime(df['date']).dt.strftime('%B, %Y')


    # Создаем ExcelWriter
    with pd.ExcelWriter('/opt/airflow/dags/kz_rates.xlsx') as writer:
        for month, group in df.groupby('month'):
            group_clean = group.drop(columns='month').copy()

            # Считаем средние курсы
            avg_usd = pd.to_numeric(group_clean['exchange_rate_USD'].str.replace(',', '.'), errors='coerce').mean()
            avg_eur = pd.to_numeric(group_clean['exchange_rate_EUR'].str.replace(',', '.'), errors='coerce').mean()

            # Создаем строку со средними
            avg_row = {
                'date': 'Среднее за месяц',
                'exchange_rate_USD': f"{avg_usd:.4f}",
                'exchange_rate_EUR': f"{avg_eur:.4f}"
            }

            # Добавляем в конец
            group_clean = pd.concat([group_clean, pd.DataFrame([avg_row])], ignore_index=True)

            # Сохраняем в лист
            group_clean.to_excel(writer, sheet_name=month, index=False)

        print('Файл Казахстана успешно сохранен')

def uz_exchange(start_date, end_date):
    data = {'date':[],
            'usd_rate':[],
           'eur_rate':[]}

    # Настройки для браузера: скрытый режим, размер окна и отключение GPU
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')  # важно в контейнере
    options.add_argument('--disable-dev-shm-usage')  # важно в контейнере
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36'
    options.add_argument(f'user-agent={user_agent}')

    driver = webdriver.Remote(
        command_executor='http://selenium:4444/wd/hub',
        options=options
    )

    url = "https://cbu.uz/ru/arkhiv-kursov-valyut/index.php"
    driver.get(url)

    # Словарь для хранения дат и курсов

    # Находим элемент текущего месяца в календаре (это span внутри div)
    month_label = driver.find_element(By.CSS_SELECTOR, 'div.xdsoft_month span')
    time.sleep(0.2)  # дать странице подвинуться
    # Кликаем по текущему месяцу через JS — открываем список месяцев
    month_label.click()

    time.sleep(1.5)

    # Проходим по найденным месяцам и кликаем по нужному
    month_list = driver.find_elements(By.CSS_SELECTOR, 'div.xdsoft_option')
    for m in month_list:
        if m.get_attribute("data-value") == str(start_date.month - 1):
            m.click()
            print(m.text)
            break

    for day in range(start_date.day, end_date.day + 1):
        try:
            calendar = driver.find_element(By.CLASS_NAME, "xdsoft_datetimepicker")
            date_cells = calendar.find_elements(By.CLASS_NAME, "xdsoft_date")

            # повторяем до 3 попыток при stale element
            attempts = 3
            while attempts > 0:
                try:
                    date_cell = next((cell for cell in calendar.find_elements(By.CLASS_NAME, "xdsoft_date")
                                      if cell.text.strip() == str(day)), None)
                    if date_cell:
                        driver.execute_script("arguments[0].scrollIntoView(true);", date_cell)
                        time.sleep(0.2)
                        driver.execute_script("arguments[0].click();", date_cell)
                        break  # клик успешен, выходим из while
                    else:
                        print(f"{day} не найден")
                        break
                except StaleElementReferenceException:
                    print(f"Stale element, retrying day {day}")
                    attempts -= 1
                    time.sleep(1)

            time.sleep(2)

            usd_rate = driver.find_element(By.XPATH,
                                           '//td[text()="USD"]/following-sibling::td[@class="text-right"]/span[@class="currency_exchange"]').text
            eur_rate = driver.find_element(By.XPATH,
                                           '//td[text()="EUR"]/following-sibling::td[@class="text-right"]/span[@class="currency_exchange"]').text
            date = datetime(start_date.year, start_date.month, day).strftime('%d.%m.%Y')

            data['usd_rate'].append(usd_rate)
            data['eur_rate'].append(eur_rate)
            data['date'].append(date)

        except Exception as e:
            print(f"Ошибка, день: {day}: {e}")
            continue
    driver.quit()

    df = pd.DataFrame(data)

    wb = Workbook()
    del wb['Sheet']  # удалим пустой лист

    # Добавим столбец с месяцем в формате "YYYY-MM"
    df['month'] = pd.to_datetime(df['date'], format="%d.%m.%Y").dt.strftime('%B, %Y')

    with pd.ExcelWriter('/opt/airflow/dags/uz_rates.xlsx') as writer:
        for month, group in df.groupby('month'):
            group_clean = group.drop(columns='month').copy()

            # Считаем средние значения
            avg_usd = pd.to_numeric(group_clean['usd_rate'].str.replace(',', '.'), errors='coerce').mean()
            avg_eur = pd.to_numeric(group_clean['eur_rate'].str.replace(',', '.'), errors='coerce').mean()

            # Строка со средними
            avg_row = {
                'date': 'Среднее за месяц',
                'usd_rate': f"{avg_usd:.4f}",
                'eur_rate': f"{avg_eur:.4f}"
            }

            # Добавляем строку
            group_clean = pd.concat([group_clean, pd.DataFrame([avg_row])], ignore_index=True)

            # Сохраняем лист
            group_clean.to_excel(writer, sheet_name=month, index=False)
        print('Файл Узбекистана успешно сохранен')

def kg_exchange(start_date, end_date):
    exchange_dict = {'date': [],
                     'exchange_rate_USD': [],
                     'exchange_rate_EUR': []}

    usd_url = f'https://www.nbkr.kg/index1.jsp?item=1562&lang=RUS&valuta_id=15&beg_day={start_date.day}&beg_month={start_date.month}&beg_year={start_date.year}&end_day={end_date.day}&end_month={end_date.month}&end_year={end_date.year}'
    eur_url = f'https://www.nbkr.kg/index1.jsp?item=1562&lang=RUS&valuta_id=20&beg_day={start_date.day}&beg_month={start_date.month}&beg_year={start_date.year}&end_day={end_date.day}&end_month={end_date.month}&end_year={end_date.year}'
    for url in [usd_url, eur_url]:
        response = requests.get(url, headers=headers)
        soap = BeautifulSoup(response.text, 'html.parser')
        data_cells = soap.find_all('td', class_='stat-center')
        rate_cells = soap.find_all('td', class_='stat-right')
        if (url == usd_url):
            for cell in rate_cells: exchange_dict['exchange_rate_USD'].append(cell.text.strip())
        else:
            for cell in rate_cells: exchange_dict['exchange_rate_EUR'].append(cell.text.strip())
            for cell in data_cells: exchange_dict['date'].append(cell.text.strip())
    df = pd.DataFrame(exchange_dict)

    wb = Workbook()
    del wb['Sheet']

    df['month'] = pd.to_datetime(df['date'], format='%d.%m.%Y').dt.strftime('%B, %Y')

    with pd.ExcelWriter('/opt/airflow/dags/kg_rates.xlsx') as writer:
        for month, group in df.groupby('month'):
            group_clean = group.drop(columns='month').copy()

            # Считаем средние значения
            avg_usd = pd.to_numeric(group_clean['exchange_rate_USD'].str.replace(',', '.'), errors='coerce').mean()
            avg_eur = pd.to_numeric(group_clean['exchange_rate_EUR'].str.replace(',', '.'), errors='coerce').mean()

            # Добавляем строку со средними
            avg_row = {
                'date': 'Среднее за месяц',
                'exchange_rate_USD': f"{avg_usd:.4f}",
                'exchange_rate_EUR': f"{avg_eur:.4f}"
            }

            group_clean = pd.concat([group_clean, pd.DataFrame([avg_row])], ignore_index=True)

            # Сохраняем на лист Excel
            group_clean.to_excel(writer, sheet_name=month, index=False)
        print('Файл Киргизии успешно сохранен')

def arm_exchange(start_date, end_date):
    start_date_url = start_date - timedelta(days=4)
    end_date_url = end_date + timedelta(days=4)

    url = f"https://www.cba.am/en/exchange-rates-archive/export?currencies%5B12%5D=EUR&currencies%5B44%5D=USD&date-from={start_date_url.day}%2F{start_date_url.month}%2F{start_date_url.year}&date-to={end_date_url.day}%2F{end_date_url.month}%2F{end_date_url.year}"

    response = requests.get(url)
    if response.status_code != 200:
        raise Exception("Не удалось скачать Excel-файл с сайта НБГ")

    # Читаем Excel из памяти
    excel_file = BytesIO(response.content)

    df_raw = pd.read_csv(excel_file, encoding='cp1251', header=None)

    df = df_raw.drop(0)
    df.columns.name = None
    df = df.reset_index(drop=True)
    df.columns = ['date', 'eur_rate', 'usd_rate']
    df = df[['date', 'usd_rate', 'eur_rate']]

    df['date'] = pd.to_datetime(df['date'], format="%d.%m.%Y")
    df = df.sort_values(by='date').reset_index(drop=True)

    full_range = pd.date_range(start=df['date'].min(), end=df['date'].max())
    df = df.set_index('date')

    # Реставрируем даты, которые отсутствуют
    df = df.reindex(full_range)

    # Заполняем пропуски предыдущим значением
    df = df.ffill()

    # Возвращаем индекс как колонку
    df = df.reset_index().rename(columns={'index': 'date'})

    df = df[(df['date'].dt.month == start_date.month) & (df['date'].dt.year == start_date.year)]

    df['date'] = df['date'].dt.strftime('%d.%m.%Y')

    wb = Workbook()

    del wb['Sheet']

    df['month'] = pd.to_datetime(df['date'], format=('%d.%m.%Y')).dt.strftime('%B, %Y')

    with pd.ExcelWriter('/opt/airflow/dags/arm_rates.xlsx') as writer:
        for month, group in df.groupby('month'):
            group_clean = group.drop(columns='month').copy()

            # Считаем средние значения
            avg_usd = pd.to_numeric(group_clean['usd_rate'].str.replace(',', '.'), errors='coerce').mean()
            avg_eur = pd.to_numeric(group_clean['eur_rate'].str.replace(',', '.'), errors='coerce').mean()

            # Добавляем строку со средними
            avg_row = {
                'date': 'Среднее за месяц',
                'usd_rate': f"{avg_usd:.4f}",
                'eur_rate': f"{avg_eur:.4f}"
            }

            group_clean = pd.concat([group_clean, pd.DataFrame([avg_row])], ignore_index=True)

            # Сохраняем лист
            group_clean.to_excel(writer, sheet_name=month, index=False)
        print('Файл Армении успешно сохранен')

def az_exchange(start_date, end_date):
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')  # важно в контейнере
    options.add_argument('--disable-dev-shm-usage')  # важно в контейнере
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36'
    options.add_argument(f'user-agent={user_agent}')

    driver = webdriver.Remote(
        command_executor='http://selenium:4444/wd/hub',
        options=options
    )

    url = "https://www.cbar.az/currency/custom?language=en"
    driver.get(url)

    # Словарь для хранения дат и курсов
    data = {'date': [],
            'usd_rate': [],
            'eur_rate': []}

    # Диапазон дат, для которых нужно получить данные
    start_date_str = start_date.strftime('%d/%m/%Y')
    end_date_str = end_date.strftime('%d/%m/%Y')

    date_from = driver.find_element(By.ID, 'currencyform-datefrom')
    date_to = driver.find_element(By.ID, 'currencyform-dateto')

    submit = driver.find_element(By.CLASS_NAME, 'table_submit')

    driver.execute_script("arguments[0].value = arguments[1];", date_from, start_date_str)
    driver.execute_script("arguments[0].dispatchEvent(new Event('change'))", date_from)
    driver.execute_script("arguments[0].value = arguments[1];", date_to, end_date_str)
    driver.execute_script("arguments[0].dispatchEvent(new Event('change'))", date_to)
    submit.click()

    rows = driver.find_elements(By.CLASS_NAME, 'table_row')
    for row in rows:
        data['date'].append(row.find_element(By.CLASS_NAME, 'valuta').text)
        data['usd_rate'].append(row.find_element(By.CLASS_NAME, 'kod').text)

    next_a = driver.find_element(By.CSS_SELECTOR, 'a.page-next')
    time.sleep(0.5)
    driver.execute_script("arguments[0].click();", next_a)
    time.sleep(3)

    rows = driver.find_elements(By.CLASS_NAME, 'table_row')
    for row in rows:
        data['date'].append(row.find_element(By.CLASS_NAME, 'valuta').text)
        data['usd_rate'].append(row.find_element(By.CLASS_NAME, 'kod').text)

    rate_button = driver.find_element(By.ID, 'currencyform-currencycode')
    rate_button_main = rate_button.find_element(By.XPATH, './following-sibling::div')
    rate_button_main.click()

    eur = driver.find_element(By.XPATH, '//li[text()="1 Euro"]')
    eur.click()
    submit = driver.find_element(By.CLASS_NAME, 'table_submit')
    # --- после нажатия на submit для евро:
    submit.click()

    # Обязательно дождись загрузки новых данных!
    time.sleep(3)

    # Теперь получи свежие элементы:
    rows = driver.find_elements(By.CLASS_NAME, 'table_row')

    for row in rows:
        data['eur_rate'].append(row.find_element(By.CLASS_NAME, 'kod').text)

    next_a = driver.find_element(By.CSS_SELECTOR, 'a.page-next')
    time.sleep(0.5)
    driver.execute_script("arguments[0].click();", next_a)
    time.sleep(3)

    # Теперь получи свежие элементы:
    rows = driver.find_elements(By.CLASS_NAME, 'table_row')

    for row in rows:
        data['eur_rate'].append(row.find_element(By.CLASS_NAME, 'kod').text)

    driver.quit()

    df = pd.DataFrame(data)

    wb = Workbook()
    del wb['Sheet']

    df['month'] = pd.to_datetime(df['date'], format='%d.%m.%Y').dt.strftime('%B, %Y')

    with pd.ExcelWriter('/opt/airflow/dags/az_rates.xlsx') as writer:
        for month, group in df.groupby('month'):
            group_clean = group.drop(columns='month').copy()

            # Считаем средние значения
            avg_usd = pd.to_numeric(group_clean['usd_rate'].str.replace(',', '.'), errors='coerce').mean()
            avg_eur = pd.to_numeric(group_clean['eur_rate'].str.replace(',', '.'), errors='coerce').mean()

            # Добавляем строку со средними
            avg_row = {
                'date': 'Среднее за месяц',
                'usd_rate': f"{avg_usd:.4f}",
                'eur_rate': f"{avg_eur:.4f}"
            }

            group_clean = pd.concat([group_clean, pd.DataFrame([avg_row])], ignore_index=True)

            # Сохраняем лист
            group_clean.to_excel(writer, sheet_name=month, index=False)
        print('Файл Азербайджана успешно сохранен')

def by_exchange(start_date, end_date):
    exchange_dict = {'date': [],
                     'usd_rate': [],
                     'eur_rate': []
                     }

    current_date = start_date

    while current_date <= end_date:
        date_str = current_date.strftime("%Y-%m-%d")
        url_usd = f"https://api.nbrb.by/exrates/rates/431?ondate={date_str}&parammode=0"
        url_eur = f"https://api.nbrb.by/exrates/rates/451?ondate={date_str}&parammode=0"
        r_usd = requests.get(url_usd)
        r_eur = requests.get(url_eur)
        exchange_dict['date'].append(current_date.strftime("%d.%m.%Y"))
        if r_usd.status_code == 200:
            exchange_dict['usd_rate'].append(r_usd.json()["Cur_OfficialRate"])
        if r_eur.status_code == 200:
            exchange_dict['eur_rate'].append(r_eur.json()["Cur_OfficialRate"])
        current_date += timedelta(1)

    df = pd.DataFrame(exchange_dict)

    wb = Workbook()
    del wb['Sheet']

    df['month'] = pd.to_datetime(df['date'], format="%d.%m.%Y").dt.strftime('%B, %Y')

    with pd.ExcelWriter('/opt/airflow/dags/by_rates.xlsx') as writer:
        for month, group in df.groupby('month'):
            group_clean = group.drop(columns='month').copy()

            # Считаем средние значения
            avg_usd = pd.to_numeric(group_clean['usd_rate'], errors='coerce').mean()
            avg_eur = pd.to_numeric(group_clean['eur_rate'], errors='coerce').mean()

            # Добавляем строку со средними
            avg_row = {
                'date': 'Среднее за месяц',
                'usd_rate': f"{avg_usd:.4f}",
                'eur_rate': f"{avg_eur:.4f}"
            }

            group_clean = pd.concat([group_clean, pd.DataFrame([avg_row])], ignore_index=True)

            # Сохраняем лист
            group_clean.to_excel(writer, sheet_name=month, index=False)
        print('Файл Беларуси успешно сохранен')

def mn_exchange(start_date, end_date):
    # Настройки для браузера: скрытый режим, размер окна и отключение GPU
    # Настройки для браузера: скрытый режим, размер окна и отключение GPU
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')  # важно в контейнере
    options.add_argument('--disable-dev-shm-usage')  # важно в контейнере
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36'
    options.add_argument(f'user-agent={user_agent}')

    driver = webdriver.Remote(
        command_executor='http://selenium:4444/wd/hub',
        options=options
    )

    url = "https://www.mongolbank.mn/en/currency-rate-movement"
    driver.get(url)

    # Словарь для хранения дат и курсов
    data = {'date': [],
            'usd_rate': [],
            'eur_rate': []}

    # Диапазон дат, для которых нужно получить данные
    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')

    date_from = driver.find_element(By.ID, 'min')
    date_from.send_keys(start_date_str)

    date_to = driver.find_element(By.ID, 'max')
    date_to.send_keys(end_date_str)

    time.sleep(7)

    submit = driver.find_element(By.ID, 'clearButton')
    driver.execute_script("arguments[0].click();", submit)

    table = driver.find_elements(By.CSS_SELECTOR, 'table.table tbody tr')

    for row in table:
        cells = row.find_elements(By.CSS_SELECTOR, 'td')
        data['date'].append(cells[1].text)
        data['usd_rate'].append(cells[2].text)
        data['eur_rate'].append(cells[3].text)

    time.sleep(4)

    next_button = driver.find_element(By.CSS_SELECTOR, 'li.paginate_button.next')

    driver.execute_script("arguments[0].scrollIntoView(true);", next_button)

    driver.execute_script("arguments[0].click();", next_button)

    time.sleep(5)

    table = driver.find_elements(By.CSS_SELECTOR, 'table.table tbody tr')

    for row in table:
        cells = row.find_elements(By.CSS_SELECTOR, 'td')
        data['date'].append(cells[1].text)
        data['usd_rate'].append(cells[2].text)
        data['eur_rate'].append(cells[3].text)

    driver.quit()

    df = pd.DataFrame(data)

    wb = Workbook()
    del wb['Sheet']  # удалим пустой лист

    df['date'] = pd.to_datetime(df['date'], format="%Y-%m-%d").dt.strftime('%d.%m.%Y')

    # Добавим столбец с месяцем в формате "YYYY-MM"
    df['month'] = pd.to_datetime(df['date'], format="%d.%m.%Y").dt.strftime('%B, %Y')

    with pd.ExcelWriter('/opt/airflow/dags/mn_rates.xlsx') as writer:
        for month, group in df.groupby('month'):
            group_clean = group.drop(columns='month').copy()

            group_clean['usd_rate'] = group_clean['usd_rate'].str.replace(',', '', regex=False)
            group_clean['eur_rate'] = group_clean['eur_rate'].str.replace(',', '', regex=False)

            avg_usd = pd.to_numeric(group_clean['usd_rate'], errors='coerce').mean()
            avg_eur = pd.to_numeric(group_clean['eur_rate'], errors='coerce').mean()

            # Добавляем строку со средними
            avg_row = {
                'date': 'Среднее за месяц',
                'usd_rate': f"{avg_usd:.4f}",
                'eur_rate': f"{avg_eur:.4f}"
            }

            group_clean = pd.concat([group_clean, pd.DataFrame([avg_row])], ignore_index=True)

            # Сохраняем лист
            group_clean.to_excel(writer, sheet_name=month, index=False)
        print('Файл Монголии успешно сохранен')
def ge_xchange():
    url = "https://nbg.gov.ge/fm/%E1%83%A1%E1%83%A2%E1%83%90%E1%83%A2%E1%83%98%E1%83%A1%E1%83%A2%E1%83%98%E1%83%99%E1%83%90/exchange_rates/eng/monthly-exchange-rateseng.xlsx?v=fyxbw"

    response = requests.get(url)
    if response.status_code != 200:
        raise Exception("Не удалось скачать Excel-файл с сайта НБГ")

    # Читаем Excel из памяти
    excel_file = BytesIO(response.content)
    df_raw = pd.read_excel(excel_file)

    # Обработка: заголовки и очистка
    currencies = df_raw.iloc[0].ffill(axis=0)
    types = df_raw.iloc[2]
    df_raw.columns = currencies.astype(str) + " - " + types.astype(str)
    df_raw = df_raw.rename(columns={df_raw.columns[0]: "Period"})
    df = df_raw[3:].copy()

    # Дата и выбор нужных колонок
    df['Period'] = pd.to_datetime(df['Period'])
    df = df[[
        'Period',
        'US Dollar - End of Month',
        'US Dollar - Monthly Average',
        'EURO - End of Month',
        'EURO - Monthly Average'
    ]]

    df = df.dropna().tail(12)
    df['Period'] = df['Period'].dt.strftime('%d.%m.%Y')

    # Сохраняем в Airflow-папку
    save_path = '/opt/airflow/dags/monthly-exchange-rateseng.xlsx'
    with pd.ExcelWriter(save_path) as writer:
        df.to_excel(writer, sheet_name='Курс за год', index=False)
        print(f'Файл успешно сохранен: {save_path}')

def pred_exchange(start_date, end_date):
    start_str = (start_date - timedelta(days=1)).strftime('%Y-%m-%d')
    end_str = (end_date + timedelta(days=1)).strftime('%Y-%m-%d')

    url = f"https://www.cbpmr.net/csv.php?vid=val&in_date={start_str}&out_date={end_str}&kod=840,978&lang=ru"

    response = requests.get(url)
    if response.status_code != 200:
        raise Exception("Не удалось скачать Excel-файл с сайта Преднестровья")

    # Читаем Excel из памяти
    excel_file = BytesIO(response.content)

    df_raw = pd.read_csv(excel_file, encoding='cp1251', header=None)

    df = df_raw.pivot_table(index=0, columns=2, values=4)
    df.columns.name = None
    df = df.reset_index()
    df.columns = ['date', 'eur_rate', 'usd_rate']
    df = df[['date', 'usd_rate', 'eur_rate']]

    df['date'] = pd.to_datetime(df['date'], format="%d.%m.%Y")
    df = df.sort_values(by='date').reset_index(drop=True)

    full_range = pd.date_range(start=df['date'].min(), end=df['date'].max())
    df = df.set_index('date')

    # Реставрируем даты, которые отсутствуют
    df = df.reindex(full_range)

    # Заполняем пропуски предыдущим значением
    df = df.ffill()

    df = df.iloc[1:-1]

    # Возвращаем индекс как колонку
    df = df.reset_index().rename(columns={'index': 'date'})

    df['date'] = df['date'].dt.strftime('%d.%m.%Y')

    df['month'] = pd.to_datetime(df['date'], format="%d.%m.%Y").dt.strftime('%B, %Y')

    wb = Workbook()
    del wb['Sheet']

    with pd.ExcelWriter('/opt/airflow/dags/pred_exchange.xlsx') as writer:
        for month, group in df.groupby('month'):
            group_clean = group.drop(columns='month').copy()
            avg_usd = pd.to_numeric(group_clean['usd_rate'], errors='coerce').mean()
            avg_eur = pd.to_numeric(group_clean['eur_rate'], errors='coerce').mean()
            group_clean = pd.concat([
                group_clean,
                pd.DataFrame([{
                    'date': 'Среднее за месяц',
                    'usd_rate': f"{avg_usd:.4f}",
                    'eur_rate': f"{avg_eur:.4f}"
                }])
            ], ignore_index=True)
            group_clean.to_excel(writer, sheet_name=month, index=False)

    print("Данные успешно сохранены.")


#kz_exchange(start_date, end_date)
# tj_exchange(start_date, end_date)
# uz_exchange(start_date, end_date)
# kg_exchange(start_date, end_date)
#arm_exchange(start_date, end_date)
#az_exchange(start_date, end_date)
#by_exchange(start_date, end_date)
#mn_exchange(start_date, end_date)
#ge_xchange()
